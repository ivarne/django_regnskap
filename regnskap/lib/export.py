# -*- coding: utf-8 -*-
"""
Collection of exportation routines
"""

from django_regnskap.regnskap.models import *
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
import datetime

class ExelYearView(object):
    """Exporter regnskapet i excel format for year"""
    
    def __init__(self, year, wb = None):
        """
        Initialize an excel export for year an using an excel file as a template
        """
        self._year = year
        self._wb = wb or openpyxl.Workbook()
        self._setWbProperties()
        #self._generateMainSheet( year, self._wb.worksheets[0] )
        del self._wb.worksheets[0]
        for prosjekt in Prosjekt.objects.all():
            sheet = self._wb.create_sheet()
            self._generateProjectOverviewSheet( prosjekt, year, sheet )
            sheet = self._wb.create_sheet()
            self._generateProjectBilagSheet( prosjekt, year, sheet )
        self._generateFullBilagSheet(year,self._wb.create_sheet())
        
    
    def _setWbProperties(self):
        # set metadata
        p = self._wb.properties
        p.creator = "Django_regnskap, av Trondheim Kristne Studentlag"
        p.last_modified_by = "Django_regnskap, av Trondheim Kristne Studentlag"
        p.title = "Regnskap for %d" % self._year
        p.description = "Autogenerert regnskap for %d\nEksportert :%Y-%m-%d %H:%M:%S".format(self._year, datetime.datetime.now())
    
    #def _generateMainSheet(self,year, sheet):
    #    sheet.title = "Oversikt"
        
    def _generateProjectOverviewSheet(self, prosjekt, year, sheet ):
        sheet.title = "Oversikt %s" % str(prosjekt)
        h = sheet.cell("A1")
        h.value = "Oversikt over Regnskapet til %s" % prosjekt
        h.style.font.size = 13
        h.style.font.bold = True
        sheet.merge_cells("A1:B1")
        
        kontos = list(Konto.objects.bilagRelated(related=prosjekt, year = year))
        #grupper kontoene
        groups = [[] for _ in range(9)]
        for konto in kontos:
            groups[konto.kontoType].append(konto)
        def kontos_list_sum(title, kontos, row, col, sign = 1, min_length = None):
            h = sheet.cell(row=row, column=col)
            h.value = title
            h.style.borders.bottom.border_style = 'thin'
            h.style.font.bold = True
            sheet.cell(row=row, column=col+1).style.borders.bottom.border_style = 'thin'
            sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
            row +=1
            for konto in kontos:
                h = sheet.cell(row = row, column = col)
                h.value = "%s %s" % (konto.nummer, konto.tittel)
                h = sheet.cell(row = row, column = col+1)
                h.value = sign*(konto.sum_kredit-konto.sum_debit)
                h.style.number_format.format_code = u'#\u00A0##0.00'
                row +=1
            if min_length:
                row += min_length - len(kontos)
            h = sheet.cell(row = row, column = col)
            h.value = "SUM"
            h.style.borders.top.border_style = 'thin'
            h.style.font.bold = True
            h = sheet.cell(row = row, column = col+1)
            column = openpyxl.cell.get_column_letter(col+2)
            h.value = "=SUM(%c%d:%c%d)" % (column,row-(min_length or len(kontos))+1,column, row)
            h.style.number_format.format_code = u'#\u00A0##0.00'
            h.style.borders.top.border_style = 'thin'
            h.style.font.bold = True
            return row+1
        row_inn = kontos_list_sum("Salgs og driftsinntekt", groups[3], 1, 0, 1)
        row_kost = kontos_list_sum("Driftskostnader", groups[4]+groups[5]+groups[6]+groups[7], row_inn+1, 0, -1)
        row_fin = kontos_list_sum("Finanskostnader og Resultat", groups[8], row_kost+1, 0, -1)
        
        sheet.cell(row = row_fin + 2, column = 0).value = u'Ikke ført resultat'
        sheet.cell(row = row_fin + 2, column = 1).value = u'=B%d-B%d-B%d' % (row_inn, row_kost, row_fin)
        
        #vis inngående balanse
        row = 0
        h = sheet.cell(row = row, column = 3)
        h.value = u"Inngående balanse %s %s"%(prosjekt, year)
        h.style.font.bold = True
        h.style.alignment.horizontal = 'center'
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row+=1
        inn_konto = Konto.objects.bilagRelated(related = prosjekt, year = year, kontoTypes=(1,2), bilagTypes=Bilag.INNGAAENDE_BALANSE)
        inn_groups = [None, [],[]]
        for konto in inn_konto:
            inn_groups[konto.kontoType].append(konto)
        min_length = max([len(a) for a in inn_groups if isinstance(a,list)])
        kontos_list_sum("Eiendeler", inn_groups[1], row, 3, -1, min_length)
        kontos_list_sum("Egenkapital og gjeld", inn_groups[2], row, 5, 1, min_length)
        row+=min_length+3
        #vis utgående balanse
        h = sheet.cell(row = row, column = 3)
        h.value = u"Utgående balanse %s %s"%(prosjekt, year)
        h.style.font.bold = True
        h.style.alignment.horizontal = 'center'
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row+=1
        min_length = max(len(groups[1]),len(groups[2]))
        kontos_list_sum("Eiendeler", groups[1], row, 3, -1, min_length)
        kontos_list_sum("Egenkapital og gjeld", groups[2], row, 5, 1, min_length)
        row+=min_length+3
        
        #set column width
        sheet.column_dimensions['A'].width = 30.
        sheet.column_dimensions['B'].width = 12.
        sheet.cell("C1")#create one element in row c so that the width can be set
        sheet.column_dimensions['C'].width = 2.
        sheet.column_dimensions['D'].width = 30.
        sheet.column_dimensions['E'].width = 12.
        sheet.column_dimensions['F'].width = 30.
        sheet.column_dimensions['G'].width = 12.
        
    def _generateProjectBilagSheet(self, prosjekt, year, sheet):
        sheet.title = "Bilag %s" % str(prosjekt)
        kontoList = list(Konto.objects.prosjekt(prosjekt))
        bilagList = Bilag.objects.prosjekt(prosjekt).filter(dato__year = year).order_by('dato')
        self._generateBilagSheet(kontoList,bilagList,year,prosjekt, sheet)
    
    def _generateFullBilagSheet(self, year, sheet):
        sheet.title = "Alle bilag"
        kontoList = list(Konto.objects.all())
        bilagList = Bilag.objects.filter(dato__year = year).order_by('bilagsnummer')
        self._generateBilagSheet(kontoList,bilagList,year,"Alle",sheet)
    
    def _generateBilagSheet(self,kontoList,bilagList,year,prosjekt,sheet):
        sheet.cell("A1").value = "Bilag"
        sheet.cell("A2").value = "#"
        sheet.cell("B1").value = "Dato"
        sheet.cell("B2").value = ""
        sheet.cell("C1").value = "Beskrivelse"
        sheet.cell("C2").value = ""
        sheet.cell("D1").value = "Hvem?"
        sheet.cell("D2").value = ""
        sheet.column_dimensions['A'].width = 8.
        sheet.column_dimensions['B'].width = 9.
        sheet.column_dimensions['C'].width = 40.
        sheet.column_dimensions['D'].width = 30.
        
        #write konto numbers and labels (first two lines)
        kontoIndex = {} # kontonummer -> column pos
        for i, konto in enumerate(kontoList):
            sheet.cell(row=0,column=i+4).value = konto.tittel
            sheet.cell(row=1,column=i+4).value = konto.nummer
            kontoIndex[konto.nummer] = i+4;
        #write all bilag
        for i, bilag in enumerate(bilagList):
            sheet.cell(row=i+2, column=0).value = "%s-%s" % (bilag.dato.year, bilag.bilagsnummer)
            c = sheet.cell(row=i+2, column=1)
            c.value = bilag.dato
            c.style.number_format.format_code = "D. MMM"
            sheet.cell(row=i+2, column=2).value = bilag.beskrivelse
            sheet.cell(row=i+2, column=3).value = unicode(bilag.external_actor or "")
            try:
                for innslag in bilag.innslag.all():
                    c = kontoIndex[innslag.konto.nummer]
                    if(innslag.isDebit):
                        sheet.cell(row=i+2, column = c).value = innslag.belop
                    else: #negativ for kredit
                        sheet.cell(row=i+2, column = c).value = -innslag.belop
            except KeyError:
                sheet.cell(row=i+2, column = 4).value = "ERROR: innslag registrert pa konto som ikke er del av dette prosjketet"
    
    def getExcelFileStream(self):
        return save_virtual_workbook(self._wb)
    
    def save(self, filename):
        return self._wb.save(filename)
    

    
    def _generateResultArk(self, year):
        ra = self._wb.worksheets[1]
        ra.cell("A1").value = "Resultatregnskap"
        ra.cell("B3").value = "Kostnader"
        ra.cell("C3").value = "Inntekter"
        intKonto =list(getKontoaggregation(year = int(year), kontoType = '3'))
        kostKonto = list(getKontoaggregation(year = int(year), kontoType = '4,5,6,7,8,9'))
        maxLen = max(len(kostKonto),len(intKonto))
        for i, konto in enumerate(kostKonto):
            ra.cell(row=i+3, column=0).value = unicode(konto)
            ra.cell(row=i+3, column=1).value = float(konto.sum_debit or 0) - float(konto.sum_kredit or 0)
        for i, konto in enumerate(intKonto):
            ra.cell(row=i+3, column=2).value = unicode(konto)
            ra.cell(row=i+3, column=3).value = float(konto.sum_kredit or 0) - float(konto.sum_debit or 0)
        ra.cell(row=maxLen+3,column=0).value = "Sum:"
        ra.cell(row=maxLen+3,column=2).value = "Sum:"
        ra.cell(row=maxLen+3,column=1).value = "=SUM(B4:B%d)" % (maxLen+3)
        ra.cell(row=maxLen+3,column=3).value = "=SUM(D4:D%d)" % (maxLen+3)
    
    def _generateBalanseArk(self, year):
        ba = self._wb.worksheets[2]
        ba.cell("A1").value = "Balanseregnskap"
        ba.cell("B3").value = "Eiendeler"
        ba.cell("C3").value = "Egenkapital og Gjeld"
        eiendelKonto =list(getKontoaggregation(year = int(year), kontoType = '1'))
        finansKonto = list(getKontoaggregation(year = int(year), kontoType = '2'))
        maxLen = max(len(eiendelKonto),len(finansKonto))
        for i, konto in enumerate(eiendelKonto):
            ba.cell(row=i+3, column=0).value = unicode(konto)
            ba.cell(row=i+3, column=1).value = float(konto.sum_debit or 0) - float(konto.sum_kredit or 0)
        for i, konto in enumerate(finansKonto):
            ba.cell(row=i+3, column=2).value = unicode(konto)
            ba.cell(row=i+3, column=3).value = float(konto.sum_kredit or 0) - float(konto.sum_debit or 0)
        ba.cell(row=maxLen+3,column=0).value = "Sum:"
        ba.cell(row=maxLen+3,column=2).value = "Sum:"
        ba.cell(row=maxLen+3,column=1).value = "=SUM(B4:B%d)" % (maxLen+3)
        ba.cell(row=maxLen+3,column=3).value = "=SUM(D4:D%d)" % (maxLen+3)